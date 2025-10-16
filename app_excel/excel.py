# -*- coding: utf-8 -*-
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import streamlit as st

# -------------------- 页面配置 --------------------
st.set_page_config(
    page_title="Excel 行范围提取工具",
    page_icon="📊",
    layout="wide"
)

# 顶部标题 & 说明
st.markdown("""
<h1 style="margin-bottom:0.2rem;">Excel 行范围提取工具 <small style="font-size:60%;">（支持 .xls / .xlsx）</small></h1>
<div style="color:#6b7280; font-size:15px; line-height:1.6;">
<ul>
<li>上传 <b>.xls</b> 或 <b>.xlsx</b> 文件</li>
<li>输入要提取的 <b>起始/结束行</b>（Excel 自然行号，从 1 开始；第 1 行通常是表头）</li>
<li>程序会遍历每个 Sheet，提取对应行段，并把“类别”列替换为 <b>Sheet 名</b></li>
<li>可选：合并相邻相同的“类别”单元格，并自动 <b>水平/垂直居中</b></li>
<li>可选：可输入结果文件的命名 <b>水平/垂直居中</b></li>
<li>点击 <b>开始提取</b> 后执行程序，执行完成后可下载结果</li>
</ul>
</div>
""", unsafe_allow_html=True)

# -------------------- 核心处理函数 --------------------
def process_excel(uploaded_file, start_row: int, end_row: int, merge_categories: bool = True) -> io.BytesIO:
    """
    uploaded_file: Streamlit 上传文件对象 (.xls / .xlsx)
    start_row, end_row: Excel 自然行号（从1开始，含两端）
    merge_categories: 是否合并相邻相同“类别”单元格并居中
    """
    file_bytes = uploaded_file.read()
    ext = uploaded_file.name.split(".")[-1].lower()

    # 选择解析器
    if ext == "xls":
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd")
    elif ext == "xlsx":
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    else:
        raise ValueError("仅支持 .xls 或 .xlsx 文件")

    iloc_start = max(int(start_row), 2) - 1     # 至少从第2行开始，避免把表头混进来
    iloc_end_exclusive = int(end_row)           # iloc stop是开区间

    all_parts = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, header=0)

        if df.shape[0] < iloc_start + 1:
            continue

        safe_end = min(iloc_end_exclusive, df.shape[0])
        if safe_end <= iloc_start:
            continue

        part = df.iloc[iloc_start:safe_end].copy()
        if "类别" in part.columns:
            part.drop(columns=["类别"], inplace=True)
        part.insert(0, "类别", sheet)
        all_parts.append(part)

    if not all_parts:
        raise ValueError("没有可用数据：请检查行号范围或源表内容。")

    result = pd.concat(all_parts, axis=0, ignore_index=True)

    # 先写到内存，再用 openpyxl 处理合并样式
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="汇总")
    buf.seek(0)

    if merge_categories:
        wb = load_workbook(buf)
        ws = wb["汇总"]

        center = Alignment(horizontal="center", vertical="center")
        cur_val, block_start = None, None
        # 从第2行开始（第1行为表头）
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v != cur_val:
                if block_start is not None and r - 1 > block_start:
                    ws.merge_cells(start_row=block_start, start_column=1, end_row=r - 1, end_column=1)
                    ws.cell(row=block_start, column=1).alignment = center
                cur_val, block_start = v, r
        # 收尾
        if block_start is not None and ws.max_row > block_start:
            ws.merge_cells(start_row=block_start, start_column=1, end_row=ws.max_row, end_column=1)
            ws.cell(row=block_start, column=1).alignment = center

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    buf.seek(0)
    return buf

# -------------------- 表单：收集输入并按按钮执行 --------------------
with st.form("extract_form", clear_on_submit=False):
    uploaded = st.file_uploader("上传 Excel 文件", type=["xls", "xlsx"])

    c1, c2, c3 = st.columns([1, 1, 1.2])
    with c1:
        start_row = st.number_input("起始行（含）", min_value=2, value=17, step=1)
    with c2:
        end_row = st.number_input("结束行（含）", min_value=2, value=22, step=1)
    with c3:
        merge_opt = st.checkbox("合并相邻相同“类别”单元格", value=True)

    file_name = st.text_input("输出文件名（无需扩展名）", value="提取结果")
    run = st.form_submit_button("🚀 开始提取", use_container_width=True)

# -------------------- 执行与结果 --------------------
if run:
    if uploaded is None:
        st.error("请先上传 Excel 文件（.xls / .xlsx）")
    elif end_row < start_row:
        st.error("结束行必须 ≥ 起始行")
    else:
        try:
            with st.spinner("正在处理，请稍候…"):
                output = process_excel(
                    uploaded_file=uploaded,
                    start_row=int(start_row),
                    end_row=int(end_row),
                    merge_categories=merge_opt
                )
            st.success("处理完成！")

            download_name = f"{file_name}.xlsx" if file_name.strip() else f"提取_{start_row}-{end_row}.xlsx"
            st.download_button(
                "⬇️ 下载处理后的 Excel",
                data=output,
                file_name=download_name,
                type="primary",
                use_container_width=True,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"处理失败：{e}")





